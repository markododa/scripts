#!/bin/python
#excel = '2/Manica_Flash.xlsx'
from openpyxl import load_workbook
lista = open('lista', 'r')
#wb = load_workbook(excel)
#ws = wb.active
kaj = {'D2:D200':'25.txt', 'F2:F200':'35.txt', 'H2:H200':'50.txt'}

def kelii(opseg, textfile, ws, pateka):
    textfile = open(textfile, 'a')
    #textfile.write(pateka+'\n')
    for row in ws.iter_rows(opseg):
        #print(row)
        for cell in row:
            if cell.value == None: 
                #textfile.write('Prazno'+'\n')
                break
            else:
                textfile.write(cell.value+'\n')
    textfile.close()

def koloni(kaj, ws, pateka):
    for x in kaj:
        kelii(x, kaj[x], ws, pateka)

def svee(lista):
    for pateka in lista:
        wb = load_workbook(pateka.strip('\n'))
        ws = wb.active
        koloni(kaj, ws, pateka)

svee(lista)
lista.close()
